from bs4 import BeautifulSoup
from httplib2 import Http
import re
from time import strftime
import csv
from openpyxl import load_workbook

date = strftime("%d-%b-%Y")

expiry = "23FEB2017"

def find_site(expiry):

	(resp_headers, content) = Http().request(("https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?segmentLink=17&instrument=OPTIDX&symbol=NIFTY&date="+ expiry),"GET", headers={'Accept':'*/*','Accept-Encoding':'gzip, deflate, sdch','Accept-Language':'en-US,en;q=0.8,en-GB;q=0.6','Cache-Control':'max-age=0','Connection':'keep-alive','Host':'www.nseindia.com','If-Modified-Since':'Thu,03 Apr 2014 11:27:44 GMT','If-None-Match':'"13309-533d45b0"','Referer':'https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?symbolCode=-10006&symbol=NIFTY&symbol=NIFTY&instrument=-&date=-&segmentLink=17&symbolCount=2&segmentLink=17','User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36'})

	soup = BeautifulSoup(content, "html.parser")
	try:
		nifty = soup.find('b',string=re.compile("NIFTY"))
		nifty = nifty.text
		print(nifty)

		section = soup.find('table', id = 'octable')	#find data table
		return section,nifty

	except:
			input("An error occured: Connection with nseindia cannot be established\nPress Enter to continue.")
			raise


def find_LTP(x, y,section):

	for tr in section.find_all('tr'):   #each row

		tds = tr.find_all(href=re.compile(".*=" + x + "\..*" + y))   #data cell

		tds = [ele.text for ele in tds]


		if tds != []:
			tds = tds[0].split()	#remove extra space before data
			#print(tds[0])
			return tds[0]




"""
function to open a .csv file, get the strike prices and expiry dates
from it, retireve LTPs from nseindia.com and store them in the same file
"""

def CSV_read_write(csvin,c):
	"""
	with open(csvin, 'r') as csvfile:

		reader = csv.DictReader(csvfile)
		ltp= []
		for row in reader:
			#read rows and append ltp
			print(row["Strike Price"])
			#print(row["Expiry"])
			#(section,nifty) = find_site(row["Expiry"])
			(section,nifty) = find_site(expiry)					#call nifty once
			ltp.append(find_LTP(row["Strike Price"],c,section))


		reader = csv.reader(csvfile)
		all = []
		csvfile.seek(0)
		row = next(reader)
		row.append(date + ", " + nifty)
		row0 = row
		#print(row)
		all.append(row)

		i = 0
		for row in reader:
			row.append(ltp[i])
			all.append(row)
			i+=1             #for item in r, item.append



		#write to file
	with open(csvin,'w') as csvoutput:

		writer = csv.writer(csvoutput)
		writer.writerows(all)
		"""
def xl_read_write(xl,cp):

	source = load_workbook(xl)
	wb = source.active
	"""
	wb['A1'] = 4
	source.save(file)
	"""
	#"""

	x=2
	y = wb.max_column +1

	while wb.cell(row=x,column=1).value != None:
		#print(wb.cell(row=x,column=2).value)
		#print(wb.cell(row=x,column=3).value)
		(section,nifty) = find_site((wb.cell(row=x,column=3).value))	#expiry dates
		ltp = find_LTP(str(wb.cell(row=x,column=2).value),cp,section)
		print(ltp)
		if ltp != None:
			wb.cell(row=x,column=y).value = float(ltp.replace(",",""))

		x+=1
	#"""
	"""
	"""
	wb.cell(row=1,column=y).value = (date + ", " + nifty)
	#print(wb.max_row)
	#print(wb.max_column)
	source.save(xl)

xl_read_write('CALL.xlsx','CE')
xl_read_write('PUT.xlsx','PE')

#CSV_read_write("CE.csv","CE")
#CSV_read_write("PE.csv","PE")
#(section,nifty) = find_site(expiry)
#print(find_LTP('8100',"CE",section))
#input("Data Saved for "+ date+ ".\nPress Enter to continue. ")
