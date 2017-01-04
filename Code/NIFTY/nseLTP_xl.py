from bs4 import BeautifulSoup
from httplib2 import Http
import re
from time import strftime
from openpyxl import Workbook,load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis


date = strftime("%d-%b-%Y")

#expiry = "23FEB2017"

def find_site(expiry):

	(resp_headers, content) = Http().request(("https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?segmentLink=17&instrument=OPTIDX&symbol=NIFTY&date="+ expiry),"GET", headers={'Accept':'*/*','Accept-Encoding':'gzip, deflate, sdch','Accept-Language':'en-US,en;q=0.8,en-GB;q=0.6','Cache-Control':'max-age=0','Connection':'keep-alive','Host':'www.nseindia.com','If-Modified-Since':'Thu,03 Apr 2014 11:27:44 GMT','If-None-Match':'"13309-533d45b0"','Referer':'https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?symbolCode=-10006&symbol=NIFTY&symbol=NIFTY&instrument=-&date=-&segmentLink=17&symbolCount=2&segmentLink=17','User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36'})

	soup = BeautifulSoup(content, "html.parser")
	try:
		nifty = soup.find('b',string=re.compile("NIFTY"))
		nifty = nifty.text
		#print(nifty)

		section = soup.find('table', id = 'octable')	#find data table
		return section,nifty

	except:
			input("An error occured: Connection with nseindia cannot be established\nPress Enter to continue.")
			raise


def find_LTP(x, y,section):

	for tr in section.find_all('tr'):   #each row

		tds = tr.find_all(href=re.compile(".*=" + x + "\..*" + y))   #data cell

		tds = [ele.text for ele in tds]


		if tds != []:
			tds = tds[0].split()	#remove extra space before data
			#print(tds[0])
			return tds[0]

"""
Open an excel sheet, get the strike prices and expiry dates
from it, retireve LTPs from nseindia.com and store them in the same sheet
"""
def xl_read_write(xl,cp):

	source = load_workbook(xl)
	wb = source.active

	x=2
	y = wb.max_column +1
	spgraph = []
	data = []
	spgraph.append('Strike Price')
	while wb.cell(row=x,column=2).value != None:
		(section,nifty) = find_site((wb.cell(row=x,column=3).value))	#expiry dates
		spgraph.append((wb.cell(row=x,column=2).value))
		ltp = find_LTP(str(wb.cell(row=x,column=2).value),cp,section)
		if ltp != None:
			wb.cell(row=x,column=y).value = float(ltp.replace(",",""))
		x+=1
	wb.cell(row=1,column=y).value = (date + ", " + nifty)

	data.append(spgraph)

	for col in wb.iter_cols(min_row=1, min_col = 5, max_col=wb.max_column, max_row=wb.max_row-1):
		sp1 = []
		for cell in col:
			#print(cell.value)
			sp1.append(cell.value)

		data.append(sp1)

	print(wb.max_row)
	print(wb.max_column)
	source.save(xl)
	return data

"""
Plot date vs. LTP for all strike prices in an excel sheet
"""

def plot_graph(rows,name):

    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)

    c2 = LineChart()
    c2.height = 12
    c2.width = 25
    c2.title = "LTP growth"
    c2.style = 7           #7,8,9
    c2.y_axis.title = "LTP"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'd-mmm'
    c2.x_axis.majorTimeUnit = "days"
    c2.x_axis.title = "Date"

    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    c2.set_categories(dates)


    x =0

    while x != ws.max_column-1:
        s2 = c2.series[x]
        s2.graphicalProperties.line.dashStyle = "sysDot"
        #s2.smooth = True # Make the line smooth
        x += 1



    ws.add_chart(c2, "A1")
    wb.save(name)


data = xl_read_write('CALL.xlsx','CE')
plot_graph(data,"CALLchart.xlsx")
data = xl_read_write('PUT.xlsx','PE')
plot_graph(data,"PUTchart.xlsx")
input("Data Saved for "+ date+ ".\nPress Enter to continue. ")
